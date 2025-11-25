import os
import xlrd
import openpyxl
from openpyxl.styles import PatternFill


def convert_xls_to_xlsx_with_colors(input_file):
    """Convert legacy XLS to XLSX while preserving background colors."""

    if not os.path.exists(input_file):
        print(f"❌ Error: File '{input_file}' not found.")
        return None

    if input_file.endswith('.xlsx'):
        return input_file

    print(f"🔄 Converting legacy file: {input_file}...")

    try:
        rb = xlrd.open_workbook(input_file, formatting_info=True)
    except Exception as e:
        print(f"❌ Error opening .xls file: {e}")
        print("   (Ensure xlrd==1.2.0 is installed)")
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_idx in range(rb.nsheets):
        r_sheet = rb.sheet_by_index(sheet_idx)
        w_sheet = wb.create_sheet(r_sheet.name)

        for row_idx in range(r_sheet.nrows):
            for col_idx in range(r_sheet.ncols):

                cell_value = r_sheet.cell_value(row_idx, col_idx)
                w_cell = w_sheet.cell(row=row_idx + 1, column=col_idx + 1)
                w_cell.value = cell_value

                xf_idx = r_sheet.cell_xf_index(row_idx, col_idx)
                xf = rb.xf_list[xf_idx]
                bgx = xf.background.pattern_colour_index
                color_tuple = rb.colour_map.get(bgx)

                if color_tuple:
                    hex_color = "{:02x}{:02x}{:02x}".format(*color_tuple).upper()

                    if hex_color not in ["FFFFFF", "000000"]:
                        fill = PatternFill(start_color=hex_color,
                                           end_color=hex_color,
                                           fill_type="solid")
                        w_cell.fill = fill

    output_file = input_file.replace(".xls", ".xlsx")
    if output_file == input_file:
        output_file += "x"

    wb.save(output_file)
    print(f"✅ Conversion complete! Created: {output_file}\n")
    return output_file

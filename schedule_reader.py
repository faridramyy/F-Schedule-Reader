import xlrd
import openpyxl
from openpyxl.styles import PatternFill
import os
import sys
from datetime import datetime, timedelta

# ==========================================
# FUNCTION 1: CONVERTER (C)
# Converts .xls to .xlsx preserving colors
# ==========================================
def convert_xls_to_xlsx_with_colors(input_file):
    if not os.path.exists(input_file):
        print(f"❌ Error: File '{input_file}' not found.")
        return None

    if input_file.endswith('.xlsx'):
        return input_file

    print(f"🔄 Converting legacy file: {input_file}...")
    
    try:
        # Open .xls with formatting enabled
        rb = xlrd.open_workbook(input_file, formatting_info=True)
    except Exception as e:
        print(f"❌ Error opening .xls file: {e}")
        print("   (Ensure you have installed xlrd==1.2.0)")
        return None

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Process each sheet
    for sheet_idx in range(rb.nsheets):
        r_sheet = rb.sheet_by_index(sheet_idx)
        w_sheet = wb.create_sheet(r_sheet.name)
        
        for row_idx in range(r_sheet.nrows):
            for col_idx in range(r_sheet.ncols):
                # Copy Value
                cell_value = r_sheet.cell_value(row_idx, col_idx)
                w_cell = w_sheet.cell(row=row_idx+1, column=col_idx+1)
                w_cell.value = cell_value
                
                # Copy Background Color
                xf_idx = r_sheet.cell_xf_index(row_idx, col_idx)
                xf = rb.xf_list[xf_idx]
                bgx = xf.background.pattern_colour_index
                
                # Map xlrd color index to RGB tuple
                color_tuple = rb.colour_map.get(bgx)
                
                if color_tuple:
                    # Convert to Hex
                    hex_color = "{:02x}{:02x}{:02x}".format(*color_tuple).upper()
                    if hex_color not in ['FFFFFF', '000000']:
                         fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                         w_cell.fill = fill

    output_file = input_file.replace('.xls', '.xlsx')
    if output_file == input_file: output_file += "x"
        
    wb.save(output_file)
    print(f"✅ Conversion complete! Created: {output_file}\n")
    return output_file

# ==========================================
# FUNCTION 2: PARSER (P)
# Reads the schedule, finds time slots, and adds Dates
# ==========================================
def analyze_schedule(file_path, target_name):
    if not file_path or not os.path.exists(file_path):
        print("❌ Error: Invalid file path for analysis.")
        return

    print(f"🔍 Scanning Schedule for: {target_name}...")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        return

    current_header_map = {} 
    current_day_str = ""     # Stores "Monday, Nov 24"
    found_any_shift = False

    # Helper to convert Excel Serial Date to String
    def excel_date_to_string(serial):
        try:
            # Excel base date is usually Dec 30, 1899
            serial = float(serial)
            if serial > 40000: # Simple check if it looks like a modern date
                dt = datetime(1899, 12, 30) + timedelta(days=serial)
                return dt.strftime("%A, %B %d, %Y")
            return None
        except:
            return None

    for row in sheet.iter_rows():
        first_cell_val = str(row[0].value).strip() if row[0].value else ""
        row_values = [str(c.value).strip() if c.value else "" for c in row]

        # --- 1. Detect Header (Time Slots) ---
        # We look for the row containing "10-11"
        if "10-11" in row_values:
            current_header_map = {}
            
            # A) Check if the first cell is a Date Serial (e.g., 45985.0)
            date_from_serial = excel_date_to_string(first_cell_val)
            
            if date_from_serial:
                current_day_str = date_from_serial
            else:
                # B) Fallback: Check if Day Name is at the end (e.g. "MON")
                possible_day = row_values[-1] if len(row_values[-1]) > 1 else "Unknown Date"
                current_day_str = possible_day

            # Map columns to times
            for cell in row:
                if cell.value and "-" in str(cell.value):
                    current_header_map[cell.column] = str(cell.value).strip()
            continue 

        # --- 2. Detect Staff Row ---
        if target_name.lower() in first_cell_val.lower():
            if not current_header_map: continue 

            shifts_found = []
            
            for cell in row:
                if cell.column in current_header_map:
                    fill = cell.fill
                    is_colored = False
                    
                    # Color Detection Logic
                    if fill.patternType == 'solid':
                        if fill.start_color.type == 'rgb':
                             if fill.start_color.rgb not in ['00000000', 'FFFFFFFF', None]:
                                 is_colored = True
                        elif fill.start_color.type == 'indexed': 
                             if fill.start_color.index != 64:
                                 is_colored = True
                    
                    if is_colored:
                        shifts_found.append(current_header_map[cell.column])

            # --- 3. Print Results ---
            if shifts_found:
                found_any_shift = True
                
                first_slot = shifts_found[0] # e.g. "10-11"
                last_slot = shifts_found[-1] # e.g. "15-16"
                
                start_time_raw = first_slot.split('-')[0]
                end_time_raw = last_slot.split('-')[1]

                def fmt_time(t):
                    t = int(float(t))
                    if t == 12: return "12:00 PM"
                    if t == 24 or t == 0: return "12:00 AM"
                    if t > 12: return f"{t-12}:00 PM"
                    return f"{t}:00 AM"
                
                final_start = fmt_time(start_time_raw)
                final_end = fmt_time(end_time_raw)

                print(f"🗓️  Event: Work Shift")
                print(f"    Date:  {current_day_str}")
                print(f"    Time:  {final_start} to {final_end}")
                # This string is optimized for Siri/Apple/Google Calendar copy-paste
                print(f"    Add:   \"Work Shift on {current_day_str} from {final_start} to {final_end}\"")
                print("   " + "-" * 40)
    
    if not found_any_shift:
        print(f"\n❌ No colored shifts found for {target_name}.")

# ==========================================
# MAIN EXECUTION
# ==========================================
if __name__ == "__main__":
    # 1. SETTINGS
    my_file = "Main Schedule  From November 24 to November 30.xls" 
    my_name = "Farid"
    
    # 2. RUN CONVERSION
    converted_file = convert_xls_to_xlsx_with_colors(my_file)

    # 3. RUN ANALYSIS
    if converted_file:
        analyze_schedule(converted_file, my_name)